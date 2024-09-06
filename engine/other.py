# -*- coding: utf-8 -*-


import openpyxl
import os
from openpyxl import Workbook
import shutil
import xlrd
import xlwt
from xlutils.copy import copy
from wama.utils import *
sep = os.sep

def calc_volume(file_path=None, excel_path=None, save_path = None, sheet:str=None):
    file_list = []
    file_name = []

    for file in os.listdir(file_path):
        file_temp = os.path.join(file_path, file)
        # 得到文件名和后缀名
        file_list.append(file_temp)
        portion = file.split("_")[0]
        file_name.append(portion)

    file_list.sort()
    file_name.sort()
    # 写入excel文件中
    count = 1
    workbook = openpyxl.load_workbook(excel_path)
    ws = workbook.active

    sheet1 = workbook.worksheets[3]
    if sheet1.title == sheet:  # 避免读取了错误的sheet
        for i in range(len(file_list)):
            #计算ROI体积
            maskFilePath = file_list[i]
            itk = sitk.ReadImage(maskFilePath)
            reader = sitk.ImageFileReader()
            reader.SetFileName(maskFilePath)
            mask = reader.Execute()
            maskArr = sitk.GetArrayFromImage(mask)  # order:z, y, x
            counts = np.sum(maskArr == 2) # 1VAT 2SAT
            spacing = mask.GetSpacing()  # order: x, y, z
            unitVol = np.prod(spacing)
            roiVol = unitVol * counts
            result1 = file_name[i]
            result2 = roiVol
            #写入
            ws.cell(row=count, column=3, value=result1)  # column指定特定的列
            ws.cell(row=count, column=4, value=result2)  # column指定特定的列
            count += 1

            print(result1, result2)
        workbook.save(save_path)
    else:
        print("输入的sheet有误")
    return

# 存入数据到excel中,一个个存
def writesimple2(path, data, index, column, sheetname='Sheet'):  # index是行数,colum是列数
    '''
    :param path: 存储地址
    :param list: 数据
    :param index: 行,从0开始
    :param column: 列，从0开始
    :param sheetname: sheet名称,默认为Sheet
    :return:
    '''
    if os.path.exists(path):
        bg = xlrd.open_workbook(path)
        sheets = bg.sheet_names() # 获取所有sheet名字
        if sheetname in sheets: # 存在该sheet,则追加数据
            Index = sheets.index(sheetname) # sheet对应的索引
            bg1 = copy(bg) # xlrd转xlwt
            writesheet = bg1.get_sheet(Index)
            writesheet.write(index, column, str(data))
            bg1.save(path)
        else:# 不存在sheet,就重新创建sheet
            bg2 = copy(bg)  # xlrd转xlwt
            writesheet = bg2.add_sheet(sheetname)
            writesheet.write(index, column, str(data))
            bg2.save(path)
    else:
        bg = xlwt.Workbook()  # 创建一个.xlsx文件,默认生成一个名为Sheet的sheet
        sheet = bg.add_sheet(sheetname)
        sheet.write(index, column, str(data))
        bg.save(path)

# 存入数据到excel中,一个个存
def writesimple(path, data, index, column, sheetname='Sheet'):  # index是行数,colum是列数
    '''
    :param path: 存储地址
    :param list: 数据
    :param index: 行
    :param column: 列
    :param sheetname: sheet名称,默认为Sheet
    :return:
    '''
    if os.path.exists(path):
        bg = openpyxl.load_workbook(path)
        sheets = bg.sheetnames
        if sheetname in sheets:
            sheet = bg[sheetname]
            sheet.cell(index+1, column+1, data)
            bg.save(path)
            bg.close()
        else:
            sheet = bg.create_sheet(sheetname)
            # sheet = bg[str(sheetname)]
            sheet.cell(index+1, column+1, data)
            bg.save(path)
            bg.close()
    else:
        bg = Workbook()  # 创建一个.xlsx文件,默认生成一个名为Sheet的sheet
        # 修改默认Sheet名为自定义sheet名
        bg1 = bg['Sheet']
        bg1.title = sheetname
        sheet = bg[sheetname]
        sheet.cell(index+1, column+1, data)
        bg.save(path)
        bg.close()

# 判断文件夹是否存在
def judgedir(path, RemoveFlag=False):
    if os.path.exists(path):
        if RemoveFlag: # 该目录已经存在的话，就删除
            shutil.rmtree(path)  # 空目录,有内容的目录都可以删
            os.makedirs(path)
        else:
            pass
    else:
        os.makedirs(path)

def log_loss_summary(loss, step, prefix=""):
    print("epoch {} | {}: {}".format(step + 1, prefix + "loss", np.mean(loss)))


# 获取img的spacing
def Getspacing(imageFilepath):
    '''
      imageFilepath:
      outimageFilepath:
      new_spacing: [n,n,n]
      new_spacing: x,y,z
      is_label: if True, using Interpolator `sitk.sitkNearestNeighbor`
    '''
    image = sitk.ReadImage(imageFilepath) #读原图
    spacing = np.array(image.GetSpacing())#读取原图spacing
    return spacing

#外扩取小图
def Array_crop(nii_path,mask_path,waikuo=30):
        subject1 = wama()
        subject1.appendImageFromNifti('IMG',nii_path) #加载图像,自定义名
        subject1.appendSementicMaskFromNifti('IMG',mask_path) #加载mask
        img_box=subject1.getBbox('IMG')
        # 谁是x谁是z与读法有关,注意检查
        xx = img_box[1] - img_box[0]
        yy = img_box[3] - img_box[2]
        zz = img_box[5] - img_box[4]
        # 读取nii的其他信息
        mask_itk = sitk.ReadImage(nii_path)
        spacing = mask_itk.GetSpacing()
        origin = mask_itk.GetOrigin()
        transfmat = mask_itk.GetDirection()
        allpath =[nii_path,mask_path]
        for path in allpath:
            mask_itk = sitk.ReadImage(path)
            mask_img = sitk.GetArrayFromImage(mask_itk)
            small_img = np.zeros([zz,xx+waikuo,yy+waikuo])
            small_img_array = mask_img[img_box[4]:img_box[5]+1,img_box[0]-int(waikuo/2):img_box[1]+int(waikuo/2)+1,img_box[2]-int(waikuo/2):img_box[3]+int(waikuo/2)+1]
            small_img = sitk.GetImageFromArray(small_img_array)
            small_img.SetSpacing(spacing)
            small_img.SetOrigin(origin)
            small_img.SetDirection(transfmat)

            if path == nii_path:
                small_mri = small_img
            else:
                small_roi = small_img
        return small_mri,small_roi

# Resamping 重采样
def Changespacing(image, outimageFilepath,new_spacing=[1.0,1.0,1.0],is_label = False):
    '''
      image:
      outimageFilepath:
      new_spacing: [n,n,n]
      new_spacing: x,y,z
      is_label: if True, using Interpolator `sitk.sitkNearestNeighbor`
    '''
    try:
        size = np.array(image.GetSize())#读取原图尺寸
        spacing = np.array(image.GetSpacing())#读取原图spacing
        new_spacing = np.array(new_spacing)
        new_size = size * spacing  / new_spacing #计算新尺寸
        new_spacing_refine = size * spacing / new_size #计算新spacing
        new_spacing_refine = [float(s) for s in new_spacing_refine]
        new_size = [int(s) for s in new_size]

        resample = sitk.ResampleImageFilter()
        resample.SetOutputDirection(image.GetDirection())
        resample.SetOutputOrigin(image.GetOrigin())
        resample.SetSize(new_size)
        resample.SetOutputSpacing(new_spacing_refine)
        if is_label:
            resample.SetInterpolator(sitk.sitkNearestNeighbor)
        else:
            # resample.SetInterpolator(sitk.sitkBSpline)
            resample.SetInterpolator(sitk.sitkLinear)
        newimage = resample.Execute(image)
        sitk.WriteImage(newimage, outimageFilepath)
    except:
        print('该数据有问题，未重采样')
        sitk.WriteImage(image, outimageFilepath)

#将3D图像resize
def resize3D(img, aimsize, order=3):
        """
        :param img: 3D array
        :param aimsize: list, one or three elements, like [256], or [256,56,56]
        :return:
        """
        _shape = img.shape
        if len(aimsize) == 1:
            aimsize = [aimsize[0] for _ in range(3)]
        if aimsize[0] is None:
            return zoom(img, (1, aimsize[1] / _shape[1], aimsize[2] / _shape[2]), order=order)  # resample for cube_size
        if aimsize[1] is None:
            return zoom(img, (aimsize[0] / _shape[0], 1, aimsize[2] / _shape[2]), order=order)  # resample for cube_size
        if aimsize[2] is None:
            return zoom(img, (aimsize[0] / _shape[0], aimsize[1] / _shape[1], 1), order=order)  # resample for cube_size
        return zoom(img, (aimsize[0] / _shape[0], aimsize[1] / _shape[1], aimsize[2] / _shape[2]),
                    order=order)  # resample for cube_size

#比较小的图像补0
def add_zero(img, target_size=[6, 6]):
        zz, xx, yy = img.shape
        dim_min_list = [int(target_size[0] / 2 - xx / 2), int(target_size[1] / 2 - yy / 2)]
        target_size_img = np.zeros([zz, target_size[0], target_size[1]])
        target_size_img[:, dim_min_list[0]:dim_min_list[0] + xx, dim_min_list[1]:dim_min_list[1] + yy] = img
        return target_size_img

def img_mask_consistent(imageFilepath,maskFilepath,new_maskFilepath):
    '''
      imageFilepath:
      outimageFilepath:
      new_spacing: [n,n,n]
      new_spacing: x,y,z
      is_label: if True, using Interpolator `sitk.sitkNearestNeighbor`
    '''
    image = sitk.ReadImage(imageFilepath) #读原图
    spacing = image.GetSpacing()
    origin = image.GetOrigin()
    transfmat = image.GetDirection()

    mask_itk = sitk.ReadImage(maskFilepath)
    mask_itk.SetSpacing(spacing)
    mask_itk.SetOrigin(origin)
    mask_itk.SetDirection(transfmat)
    sitk.WriteImage(mask_itk, new_maskFilepath)

def findindex(vinSets):
    '''找到列表中相同的元素，并找到其索引'''
    list_vin = []
    for i in vinSets:
        address_index = [x for x in range(len(vinSets)) if vinSets[x] == i]
        list_vin.append([i, address_index])
    dict_address = dict(list_vin)
    return dict_address