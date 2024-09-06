from openpyxl import Workbook
from openpyxl import load_workbook
from sklearn.model_selection import KFold, StratifiedKFold
from sklearn.svm import SVC
from sklearn.model_selection import GridSearchCV
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn import preprocessing
from sklearn.linear_model import LassoCV
from sklearn.linear_model import Lasso
from sklearn import feature_selection
from other import judgedir
import sklearn.metrics as metrics
import matplotlib.pyplot as plt
import h5py
import time
import os
import gc
import shutil
import warnings
import pandas as pd
import random
import numpy as np
import pickle
import copy
import torch
import other
from imblearn.over_sampling import SVMSMOTE, SMOTE

warnings.filterwarnings('ignore')
sep = os.sep
filesep = sep

def save_bestcoff(save_path, bestcoff, type):
    # Write weights to txt file
    with open(save_path, "w") as f:
        f.write(f"Type: {type}\n")
        for i, coef in enumerate(bestcoff):
            f.write(f"Feature {i+1}: {coef}\n")

    print("Model Coefficients (Weights):", bestcoff)
    print(f"Model weights saved to {output_file_path}")

class Classifier():
    '''
     :param pkl_file:Patient information pkl address
     :param savepath:Save file address
     :param inner_radio_feature_path:Internal radiomics feature address
     :param inner_dsfr_feature_path:Internal dsfr feature address
     :param out_radio_feature_path:External radiomics feature address
     :param out_dsfr_feature_path:External dsfr feature address
     :param feature_excel_path:Feature name excel address
    '''
    def __init__(self, pkl_file, savepath,radio_feature_path_0,dsfr_feature_path_0,radio_feature_path_6,dsfr_feature_path_6,task,data_name,feature_excel_path=None):
        super().__init__()
        # ======================================Path addresses=============================================
        self.pkl_file = pkl_file
        self.savepath = savepath
        self.radio_feature_path_0 = radio_feature_path_0
        self.dsfr_feature_path_0 = dsfr_feature_path_0
        self.radio_feature_path_6 = radio_feature_path_6
        self.dsfr_feature_path_6 = dsfr_feature_path_6
        self.feature_excel_path = feature_excel_path
        self.task = task
        # ======================================Other parameters=============================================
        self.data_name = data_name
        self.test_outside = True  # Whether there is an external test set
        self.holdout = False  # Whether to leave-one-out validation
        self.k_fold = 2  # Number of cross-validation folds, with balanced!! cross-validation folds
        self.N = 5000  # Run N repeated experiments
        self.show_grid = False  # Whether to print grid search details
        self.distill_mode = 'mean'  # Strategy for integrating patient blocks into a vector, max, mean or maxmean
        # Whether to use a certain classifier
        #todo
        self.paint = True # Whether to draw lasso selection graph
        self.UseSVM = False # Whether to use SVM
        self.UseLR = True # Whether to use LR
        self.UseRF = False  # Whether to use RF
        self.Smo = True # Whether to use smo resampling
        self.SvmSmo = False
        self.mutualinfo = True # Whether to use mutual information to filter features

    # Classifier parameters
    def class_para(self):
        # Determine svm parameter grid
        self.svm_kernel = ['linear', 'poly']  # Linear takes too long to run
        self.svm_C = [1e-8, 1e-6, 1e-5, 1e-3, 1e-2, 1e-1, 1, 2, 4, 8, 16, 100, 200]
        self.svm_Gamma = [1e-8, 1e-6, 1e-5, 1e-3, 1e-2, 1e-1]
        self.svm_classwight = [{0: 0.55, 1: 1}, {0: 0.6, 1: 1}, {0: 0.65, 1: 1}, {0: 0.7, 1: 1}, {0: 0.8, 1: 1},
                               {0: 1, 1: 1}]
        self.svm_max_iter = [2, 4, 10, 20, 25, 30, 40, 60, 140, 200, 300, -1]
        self.svm_grid_search_param = [
            {"kernel": self.svm_kernel, "C": self.svm_C, "gamma": self.svm_Gamma, "class_weight": self.svm_classwight,
             "max_iter": self.svm_max_iter}]


        # Modify parameters here
        # Determine logistics parameter grid
        self.lr_penalty = ['l2']
        self.lr_C = [0.0001,0.001, 0.01, 0.05, 0.1, 0.3, 0.4, 0.6, 0.8, 1.0] # best 0.01
        self.lr_max_iter = [1,10, 20, 40, 60, 80, 100, 150, 200, 400] # best 10
        self.lr_tol = [1e-9, 1e-8,1e-7, 1e-5, 1e-4, 1e-3, 1e-2] # best 1e-5
        self.lr_grid_search_param = [\
            {"penalty": self.lr_penalty, "C": self.lr_C, "max_iter": self.lr_max_iter, "tol": self.lr_tol}]

        # Determine rf parameter grid
        self.rf_n_estimators = [101, 151, 201, 251, 301, 351, 401, 451, 501, 551, 800]
        self.rf_grid_search_param = [{"n_estimators": self.rf_n_estimators}]

    # Shuffle in any axis
    def array_shuffle(self, x, axis=0):
        new_index = list(range(x.shape[axis]))
        random.shuffle(new_index)
        x_new = np.transpose(x, ([axis] + [i for i in list(range(len(x.shape))) if i is not axis]))
        x_new = x_new[new_index][:]
        new_dim = list(np.array(range(axis)) + 1) + [0] + list(np.array(range(len(x.shape) - axis - 1)) + axis + 1)
        x_new = np.transpose(x_new, tuple(new_dim))
        return x_new

    # Data preprocessing　
    def processing(self, train_Data=[], test_Data=[], processig_Index='StandardScaler'):
        if processig_Index == 'MinMaxScaler':
            min_max_scaler = preprocessing.MinMaxScaler()
            processing_train_data = min_max_scaler.fit_transform(train_Data)
            # Apply the scale parameters obtained above to the test data
            processing_test_data = min_max_scaler.transform(test_Data)
            # The scaler's properties can be viewed using the following method
            min_max_scaler.scale_
            min_max_scaler.min_
        elif processig_Index == 'StandardScaler':
            StandardScaler = preprocessing.StandardScaler()
            processing_train_data = StandardScaler.fit_transform(train_Data)
            # Apply the scale parameters obtained above to the test data
            processing_test_data = StandardScaler.transform(test_Data)
        else:
            print('try another processing method')
        return processing_train_data, processing_test_data

    # Convert dictionary data to matrix data
    def dict2np(self, data, distill_mode):
        """
        #Organize data to patient units and store patient features in a dictionary
        :param data: Dictionary containing all patient omics, semantic features and labels
        :param distill_mode: Three feature integration strategies, max, mean, maxmean
        :return: np form of all id, label, dsfr_feature, radio_feature
        """
        id_all = []
        all_key = data.keys()  # data.keys is the patient ID
        data_dict = {}
        for id in all_key:
            data_dict[id] = {'dsfrfeature': data[id]['dsfr_feature'], \
                             'radiofeature': data[id]['radio_feature'], \
                             'label': data[id]['label']}

        # Integrate features of each patient in advance (take mean or max, or mean + max titu1996 method) =======================================
        feature = []
        radiofeature = []
        label = []
        for idd in all_key:
            if distill_mode == 'max':
                feature.append(list(np.max(data_dict[idd]['dsfrfeature'], axis=0)))  # column
            elif distill_mode == 'mean':
                feature.append(list(np.mean(data_dict[idd]['dsfrfeature'], axis=0)))
            elif distill_mode == 'maxmean':
                feature.append(list(0.5 * (np.mean(data_dict[idd]['dsfrfeature'], axis=0) +
                                           np.max(data_dict[idd]['dsfrfeature'], axis=0))))
            else:
                raise Exception("distill_mode must be one of them (max,mean,maxmean) ")
            radiofeature.append(list(data_dict[idd]['radiofeature'])) # Omics features not needed
            id_all.append(idd)
            label.append(data_dict[idd]['label'])

        # For features
        dsfr_feature = np.array(feature)
        radio_feature = np.array(radiofeature)
        label = np.array(label)
        id_all = np.array(id_all)
        return [id_all, label, dsfr_feature, radio_feature]

    # Save and draw feature weight map
    def weight_map(self,Index_coel,savepath,feature_title_excel=None,alpha=None):
        '''
        :param feature_title_excel: Excel table saving feature names
        :param Index_coel: Index table after lasso screening
        :param alpha: Find the best alpha and save multiple graphs
        :return:
        '''
        # Read in radiomics feature names
        if feature_title_excel!= None: # If there are feature names, such as radiomics, import the address
            df1 = pd.read_excel(feature_title_excel) # Open the feature name excel
            data_feature_name = list(df1.columns.values) # Convert feature names to list
        else: # Deep learning has no feature names, directly use the number position instead
            data_feature_name = list(range(len(Index_coel)))
        place = np.nonzero(Index_coel) # Get the positions of non-zero elements of coel
        index = [] # Empty feature name list
        for i in place[0]:
            feature_name = data_feature_name[i] # Get the filtered feature name
            index.append(feature_name)  # Store in list
        x_values = np.arange(len(index)) # The x-axis of the image is the feature name, in np format
        y_values = Index_coel[Index_coel!= 0] # The y-axis of the image is the weight
        plt.figure(figsize=(15, 10)) # Image size
        plt.bar(x_values, y_values  # bar plot
               , color='lightblue'
               , edgecolor='black'
               , alpha=0.8  # Opacity
                )
        plt.xticks(x_values, index
                  , rotation=45
                  , ha='right'
                  , va='top'
                   )
        plt.xlabel('feature', fontsize=15, fontweight='bold')
        plt.ylabel('weight', fontsize=15, fontweight='bold')
        plt.tight_layout()  # Prevent incomplete text display
        if alpha!= None:
            # Create if the path does not exist
            if not os.path.exists(os.path.join(savepath, 'Lasso筛选权重图')):
                os.mkdir(os.path.join(savepath, 'Lasso筛选权重图'))
            # Save the weight map
            plt.savefig(os.path.join(savepath, 'Lasso筛选权重图',str(alpha)+'_feature_weight.png'))
        else:
            # Create if the path does not exist
            if not os.path.exists(os.path.join(savepath, 'Lassocv筛选权重图')):
                os.mkdir(os.path.join(savepath, 'Lassocv筛选权重图'))
            # Save the weight map
            plt.savefig(os.path.join(savepath, 'Lassocv筛选权重图','feature_weight.png'))

    # lasso algorithm
    def lasso(self, Data, label, savepath, alpha, paint=True):
        lasso_name = 'lasso_' + str(alpha) + '.png'
        model_lasso = Lasso(alpha=alpha, max_iter=500).fit(Data, label)
        coef_to_return = model_lasso.coef_
        intercept_to_return = model_lasso.intercept_

        if not paint:
            return coef_to_return, intercept_to_return
        else:
            print("alpha: ", alpha)
            StandardScaler = preprocessing.MinMaxScaler()  # Normalization
            Data = StandardScaler.fit_transform(Data)  # Find the maximum and minimum values of each dimension and save them, and normalize data_train

            alpha = np.arange(alpha, 0.1, 0.0003)

            plt.figure(figsize=(6, 6), dpi=300)
            coeff_data = []
            loss = []
            for ln in alpha:
                lasso_model = Lasso(alpha=ln)
                lasso = lasso_model.fit(Data, label)
                coef = lasso.coef_
                coeff_data.append(coef)
            coeff_data = np.array(coeff_data)
            print(np.shape(coeff_data))
            for feature_num in range(coeff_data.shape[1]):
                feature_coeff = coeff_data[:, feature_num]
                plt.plot(np.log10(alpha), feature_coeff, linewidth=1)

            plt.ylim([-0.2, 0.2])
            plt.xlim([np.log10(alpha)[0], np.log10(0.1)])  #10 to the power of -0.1
            plt.xlabel('log 10 (alpha)', fontdict={'family': 'Times New Roman', 'size': 15})
            plt.ylabel('coefficients', fontdict={'family': 'Times New Roman', 'size': 15})
            plt.savefig(os.path.join(savepath, lasso_name), dpi=1000)
        return coef_to_return

    # lassocv algorithm
    def lassocv(self,Data, label):
        alphas = np.logspace(-10, 1, 512)  # 45
        model_lassoCV = LassoCV(alphas=alphas, cv=10, max_iter=1000000).fit(Data, label)
        coef = model_lassoCV.coef_
        return coef

    # Use mutual information method to filter features
    def fitness_feature_choose(self, Data, Data_valid, label, SelectPercentile_fc=2, pers_chi=30,
                           test_outside=True,coel_index=None):
        ''' f_classif : ANOVA F-value between label/feature for classification tasks. F-value of variance analysis between label/feature for classification tasks.
            mutual_info_classif : Mutual information for a discrete target. Mutual information for discrete targets.
            chi2 : Chi-squared stats of non-negative features for classification tasks. Chi-square statistics of non-negative features for classification tasks.
            f_regression : F-value between label/feature for regression tasks. F-value between label/feature for regression tasks.
            mutual_info_regression : Mutual information for a continuous target. Mutual information for continuous targets.
            SelectKBest : Select features based on the k highest scores. Select features based on the k highest scores.
            SelectFpr : Select features based on a false positive rate Random_Or_Disuse. Select features based on false positive rate test.
            SelectFdr : Select features based on an estimated false discovery rate. Select features based on estimated false discovery rate.
            SelectFwe : Select features based on family-wise error rate. Select features based on family-wise error rate.
            GenericUnivariateSelect : Univariate feature selector with configurablemode. Univariate feature selector with configurable mode.
        '''
        pers_chi = pers_chi  # Percentage of features remaining after filtering
        if SelectPercentile_fc == 1:
            function_choose = feature_selection.f_classif  # Filtering method
        elif SelectPercentile_fc == 2:
            function_choose = feature_selection.mutual_info_classif
        elif SelectPercentile_fc == 3:
            function_choose = feature_selection.chi2
        elif SelectPercentile_fc == 4:
            function_choose = feature_selection.f_regression
        elif SelectPercentile_fc == 5:
            function_choose = feature_selection.mutual_info_regression

        fs = feature_selection.SelectKBest(function_choose, k=pers_chi)
        feature = fs.fit_transform(Data, label)
        save_index = fs.get_support(indices=True)
        after_mutual = []

        for i in range(len(save_index)):
            after_mutual.append(coel_index[save_index[i]])
        print('final index: ', after_mutual)

        if test_outside:
            feature_valid = fs.transform(Data_valid)

        print('final feature num: ', feature.shape[1])
        return feature, feature_valid, save_index

        # Select features

    def choose_feature(self, feature, savepath, feature_out=None, uselasso=False, alpha=0.0008, \
                       uselassocv=False, radio_dsfr=None):
        # ===============================================Parameter settings =================================================================
        # Default no external data
        test_outside = False
        if feature_out is not None:
            test_outside = True
        fl = 1  # Number of selected features
        # lasso class
        uselasso = uselasso
        uselassocv = uselassocv
        alpha = alpha
        print('select features first')
        if not os.path.exists(savepath):
            os.mkdir(savepath)
        # Select features using lasso
        if uselasso:
            if (self.paint):
                coel = self.lasso(feature, self.label, savepath, alpha)
            else:
                coel = self.lasso(feature, self.label, alpha)
            fl = 0
            coel_index = []
            for i in range(len(coel)):
                if coel[i] != 0:
                    fl = fl + 1
                    coel_index.append(
                        i)  # Get the index of features in feature that meet the conditions corresponding to the original h5 file.
            print('lasso significant feature num:', fl)
            print('lasso significant feature index:', coel_index)
            feature = feature[:, coel != 0]
            self.weight_map(coel, savepath, alpha=alpha)
            if test_outside:
                feature_out = feature_out[:, coel != 0]
            if (self.mutualinfo):
                print('Perform mutual information feature selection')
                if (radio_dsfr == 'radio'):
                    feature, feature_out, save_index = \
                        self.fitness_feature_choose(Data=feature, Data_valid=feature_out,
                                                    label=self.label, pers_chi=25, coel_index=coel_index)
                else:
                    feature, feature_out, save_index = \
                        self.fitness_feature_choose(Data=feature, Data_valid=feature_out,
                                                    label=self.label, pers_chi=23, coel_index=coel_index)
                print('Number of {} features after mutual information: {}'.format(radio_dsfr, feature.shape[1]))
                y = np.where(coel != 0)
                temp = np.zeros(1130)
                for i in range(len(save_index)):
                    temp[y[0][save_index[i]]] = coel[y[0][save_index[i]]]
                self.weight_map(temp, savepath, alpha=alpha + 1)  # alpha+1 indicates another graph

                max_weight = temp[temp != 0]
                print('max_weight:', max_weight)

        if uselassocv:
            coel = self.lassocv(feature, self.label)
            fl = 0
            for i in range(len(coel)):
                if coel[i] != 0:
                    fl = fl + 1
            print('lassocv significant feature num:', fl)
            feature = feature[:, coel != 0]
            self.weight_map(coel, savepath, self.feature_excel_path)
            if test_outside:
                feature_out = feature_out[:, coel != 0]

        return feature, feature_out, fl

        # Sigmoid function

    def sigmoid_y(self, x, thresold=0.5):
        if x < thresold:
            x = 0
        else:
            x = 1
        return x

        # Calculate metrics

    def getAccSenSpcAuc(self, label, pre, pre_bestthresold=None):
        """
        Only applicable for binary classification
        :param label:01 label
        :param pre:Probability of belonging to class 1
        :param pre_bestthresold:Can manually set the threshold, otherwise return the threshold at the maximum Youden index on the ROC curve.
        :return:
        """
        final_true_label = label
        final_pred_value = pre
        patient_num = len(final_true_label)

        # Calculate auc and calculate the best threshold
        if (sum(final_true_label) == patient_num) or (sum(final_true_label) == 0):
            Aucc = 0
            print('only one class')
        else:
            Aucc = metrics.roc_auc_score(final_true_label, final_pred_value)
            # print('AUC', Aucc)

        # Calculate the best threshold
        fpr, tpr, thresholds = metrics.roc_curve(final_true_label, final_pred_value)
        # Calculate the Youden index
        Youden_index = tpr + (1 - fpr)
        best_thresold = thresholds[Youden_index == np.max(Youden_index)][0]

        # Have no idea about that threshold is bigger than 1 sometimes
        # maybe can find in https://github.com/scikit-learn/scikit-learn/commit/4d9a67f77787ffe9955187865f9b95e19286f069
        # or https://github.com/scikit-learn/scikit-learn/issues/3097
        if best_thresold > 1:
            best_thresold = 0.5

        # If there is a preset threshold, use the preset threshold to calculate acc, sen, spc
        if pre_bestthresold is not None:
            best_thresold = pre_bestthresold

        # Calculate the final metrics based on the final list
        tp = 0
        tn = 0
        fp = 0
        fn = 0

        for nn in range(patient_num):
            t_label = final_true_label[nn]  # true label
            p_value = final_pred_value[nn]

            p_label = self.sigmoid_y(p_value, best_thresold)

            if (t_label == 1) and (t_label == p_label):
                tp = tp + 1  # True positive
            elif (t_label == 0) and (t_label == p_label):
                tn = tn + 1  # True negative
            elif (t_label == 1) and (p_label == 0):
                fn = fn + 1  # False negative
            elif (t_label == 0) and (p_label == 1):
                fp = fp + 1  # False positive

        Sensitivity = tp / ((tp + fn) + (1e-16))
        Specificity = tn / ((tn + fp) + (1e-16))
        Accuracy = (tp + tn) / ((tp + tn + fp + fn) + (1e-16))

        return [Accuracy, Sensitivity, Specificity, Aucc, best_thresold]

        # Store data in excel

    def writesimple(self, path, list, index, column,
                    sheetname='Sheet1'):  # index is row number, column is column number
        '''
        :param path: Storage address
        :param list: Data
        :param index: Row
        :param column: Column
        :param sheetname: Sheet name, default is 'Sheet1'
        :return:
        '''
        if os.path.exists(path):
            bg = load_workbook(path)
            sheets = bg.sheetnames
            if sheetname in sheets:
                sheet = bg[str(sheetname)]
                sheet.cell(index, column, list)
                bg.save(path)
                bg.close()
            else:
                sheet = bg.create_sheet(str(sheetname))
                sheet = bg[str(sheetname)]
                sheet.cell(index, column, list)
                bg.save(path)
                bg.close()
        else:
            bg1 = Workbook()
            bg1.active
            bg1.save(path)
            bg1.close()
            bg = load_workbook(path)
            sheet = bg.create_sheet(str(sheetname))
            # sheet = bg[str(sheetname)]
            sheet.cell(index, column, list)
            bg.save(path)
            bg.close()

    # Save grid search best optimization parameters
    def save_param(self, path, time, modelname, classifiername):
        '''
        :param path: Storage path
        :param time: Round
        :param modelname: Model name, generally clf
        :param classifiername: Classifier name such as SVM
        :return:
        '''
        tmp = modelname.best_params_.keys()  # Best parameter dictionary
        if time == 1:
            self.writesimple(path, 'turn', 1, 1, sheetname=classifiername)  # Write header on the first line
        self.writesimple(path, str(time), time + 1, 1, sheetname=classifiername)
        flag_save = 1
        for a in tmp:
            flag_save = flag_save + 1
            if time == 1:
                self.writesimple(path, str(a), 1, flag_save,
                                 sheetname=classifiername)  # Write header on the first line
            self.writesimple(path, str(modelname.best_params_[a]), time + 1, flag_save, sheetname=classifiername)

    # Run main program
    def run(self, show_type, choose_alpha_type, choose_type):
        '''
        :param show_type: Select output result 1. dsfr 2. radiomics 3. dsfr + radiomics scores added
        :param choose_alpha_type: Select mode (lasso feature selection) 1. Find the best alpha 2. Fixed alpha
        :param choose_type: Select running mode 1. Train/Test 2. Recurrent
        '''

        random.seed(6666)  # Set random number seed
        # Initialize parameters
        Inner_PATIENTS_Label = {}
        Out_PATIENTS_Label = {}
        alldata_label = pickle.load(self.pkl_file)  # Read file label file
        in_p_num = 0  # Calculate the proportion of '1' inside
        out_p_num = 0  # Calculate the proportion of '1' in the test set
        # Read the content of the pkl file
        for pth_id in alldata_label.keys():
            # pth_id = str(pth_id[1:])
            if str(alldata_label[pth_id]['Use']) == "1":
                if str(alldata_label[pth_id]['Type']) == 'train':
                    Inner_PATIENTS_Label[pth_id] = alldata_label[pth_id]['Labelall']  # Training set ID and label
                    if Inner_PATIENTS_Label[pth_id] == 1:  # Calculate the proportion of '1' inside
                        in_p_num = in_p_num + 1
                else:
                    Out_PATIENTS_Label[pth_id] = alldata_label[pth_id]['Labelall']  # External test set ID and label
                    if Out_PATIENTS_Label[pth_id] == 1:  # Calculate the proportion of '1' in the external test set
                        out_p_num = out_p_num + 1
        # Create result save folder
        self.judgedir(self.savepath)  # Create save folder
        # os.startfile(self.savepath)  # Pop up save folder, too lazy to find

    # ======================================Read data=============================================
        # Read data step1: Read internal data
        In_data_6 = {}
        for pth_id in Inner_PATIENTS_Label.keys():
            portion = pth_id.split("_")[0]
            # Created dictionary In_data: In_data[ id{'dsfr_feature','radio_feature','label'} ]
            if 'dsfr' in self.data_name:
                # Semantic feature 6
                H5_file_dsfr = h5py.File(os.path.join(dsfr_feature_path_6, str(pth_id) + '.h5'), 'r')  # Open.h5 file
                dsfr_feature_6 = H5_file_dsfr['f_values'][:]  # Read semantic feature from h5
                H5_file_dsfr.close()  # Close.h5 file

            if 'radio' in self.data_name:
                H5_file_radio = h5py.File(os.path.join(radio_feature_path_6, str(pth_id) + '.h5'),
                                          'r')  # Read file label file
                radio_feature_6 = H5_file_radio['f_values'][:][0]
                H5_file_radio.close()  # Close.h5 file

            if len(self.data_name) == 0:
                raise Exception("No data!")
            if len(self.data_name) == 1:  # If there is only one modality, then two variables save the same modality
                try:
                    dsfr_feature_6 = radio_feature_6
                    radio_feature_6 = dsfr_feature_6
                except:
                    pass

            # Create dictionary 6
            In_data_6[pth_id] = {}  # Create an empty dictionary
            In_data_6[pth_id][
                'dsfr_feature'] = dsfr_feature_6  # Put the semantic feature of this patient into the dictionary
            In_data_6[pth_id][
                'radio_feature'] = radio_feature_6  # Put the omics feature of this patient into the dictionary
            In_data_6[pth_id]['label'] = Inner_PATIENTS_Label[pth_id]  # Patient corresponding label

        # Read data step2: Read external validation data (if any)
        if self.test_outside:

            Out_data_6 = {}
            for pth_id in Out_PATIENTS_Label.keys():
                # Semantic feature 6
                H5_file = h5py.File(os.path.join(dsfr_feature_path_6, str(pth_id) + '.h5'), 'r')
                out_dsfr_feature_6 = H5_file['f_values'][:]  # Read semantic feature from h5
                H5_file.close()  # Close.h5 file

                # Omics feature 6 #todo
                H5_file = h5py.File(os.path.join(radio_feature_path_6, str(pth_id) + '.h5'),
                                    'r')  # Read file label file
                out_radio_feature_6 = H5_file['f_values'][:][0]  # Read semantic feature from h5
                H5_file.close()  # Close.h5 file

                # Create dictionary 6
                Out_data_6[pth_id] = {}  # Create an empty dictionary
                Out_data_6[pth_id][
                    'dsfr_feature'] = out_dsfr_feature_6  # Put the semantic feature of this patient into the dictionary
                Out_data_6[pth_id][
                    'radio_feature'] = out_radio_feature_6  # Put the omics feature of this patient into the dictionary
                Out_data_6[pth_id]['label'] = Out_PATIENTS_Label[pth_id]

        # ====================================== Organize data =============================================
        # step1 Organize internal data
        # 6
        id_all, self.label, dsfrfeature_6, radiofeature_6 = self.dict2np(In_data_6, self.distill_mode)
        print('{} in data num    0:1=={}:{}'.format(len(id_all), int(len(Inner_PATIENTS_Label) - in_p_num),
                                                    in_p_num))
        # step2 Organize external data (if any)
        if self.test_outside:
            # 6
            id_all_out, self.label_out, dsfrfeature_out_6, radiofeature_out_6 = self.dict2np(Out_data_6,
                                                                                             self.distill_mode)
            print('{} out data num    0:1=={}:{}'.format(len(id_all_out),
                                                         int(len(Out_PATIENTS_Label) - out_p_num), out_p_num))

        # Check features
        for x in range(dsfrfeature_out_6.shape[0]):
            for y in range(dsfrfeature_out_6.shape[1]):
                if type(dsfrfeature_out_6[x][y]) == np.ndarray:
                    dsfrfeature_out_6[x][y] = dsfrfeature_out_6[x][y].ravel()[0].astype(
                        'float64')  # Flatten to one dimension and take 0
        dsfrfeature_out_6 = dsfrfeature_out_6.astype('float64')
        # ====================================== Normalization =============================================
        # Normalize external data first, otherwise after internal data is covered, the maximum and minimum values will change. 6
        if self.test_outside:  # Normalize according to internal data
            radiofeature_out_6 = (radiofeature_out_6 - radiofeature_6.min(axis=0)) / (
                    radiofeature_6.max(axis=0) - radiofeature_6.min(axis=0) + 1e-12)
            dsfrfeature_out_6 = (dsfrfeature_out_6 - dsfrfeature_6.min(axis=0)) / (
                    dsfrfeature_6.max(axis=0) - dsfrfeature_6.min(axis=0) + 1e-12)
        # Normalize internal features
        dsfrfeature_6 = (dsfrfeature_6 - dsfrfeature_6.min(axis=0)) / (
                dsfrfeature_6.max(axis=0) - dsfrfeature_6.min(axis=0) + 1e-12)
        radiofeature_6 = (radiofeature_6 - radiofeature_6.min(axis=0)) / (
                radiofeature_6.max(axis=0) - radiofeature_6.min(axis=0) + 1e-12)

        # ====================================== Big loop starts =============================================
        flag_1 = 0
        for iii in range(self.N):
            gc.collect()  # Release memory
            flag_1 = flag_1 + 1
            print('Round:', flag_1, '/', self.N, '======================================================')
            radioflcv1 = 1
            dsfrfl = 1
            tfnfl = 1
            prob_svm_final = []
            prob_lr_final = []
            prob_rf_final = []

            prob_svm_out_final = []
            prob_lr_out_final = []
            prob_rf_out_final = []
            final_label_CV = []

            # Splitting
            num_subjects = id_all.shape[0]  # Not only use 0 features, need to be modified, temporarily like this
            if choose_type == 1:  # Mode 1 is training/testing mode
                print("Current mode 2: Train/Test!")
                cvrdstate = random.randint(1, 9999)  # Splitting random number
                print('cvrdstate', cvrdstate)
                grid_rdstate = random.randint(1, 9999)  # Classifier random number
                print('grid_rdstate', grid_rdstate)
                # Leave-one-out method
                if self.holdout:
                    sfolder = KFold(n_splits=num_subjects, random_state=cvrdstate, shuffle=True)
                else:
                    sfolder = StratifiedKFold(n_splits=self.k_fold, random_state=cvrdstate, shuffle=True)
            elif choose_type == 2:  # Mode 2 is recurrence mode
                print("Current mode is Recurrent!")
                cvrdstate = int(input("Enter cvrdstate:"))  # Enter fixed splitting random number
                grid_rdstate = int(input("Enter grid_rdstate:"))  # Enter fixed classifier random number
                self.N = 1
            else:
                raise Exception("The Choose_type must be set to '1' or '2'")

            # Run experiments by task
            for i in range(len(self.task)):
                gc.collect()  # Release memory
                task = self.task[i]
                final_savepath = os.path.join(self.savepath, task)
                if not os.path.exists(final_savepath):
                    os.mkdir(final_savepath)
                print("======================TASK={taskname}======================".format(taskname=task))
                if task == '6':
                    radiofeature = radiofeature_6
                    dsfrfeature = dsfrfeature_6
                    if self.test_outside:
                        radiofeature_out = radiofeature_out_6
                        dsfrfeature_out = dsfrfeature_out_6
                    if show_type == 3:
                        # Omics feature selection
                        alpha = 0.0021  # fixed
                        print('Omics feature selection, alpha = {}'.format(alpha))
                        inner_radio_feature_choose, out_radio_feature_choose, radioflcv1 = \
                            self.choose_feature(radiofeature, feature_out=radiofeature_out, \
                                                savepath=final_savepath + r"\组学", uselasso=True, alpha=alpha,
                                                radio_dsfr='radio')

                        # Semantic segmentation feature selection
                        if choose_alpha_type == 1:  # Find the maximum alpha, increase alpha with each cycle
                            if flag_1 == 1:
                                print("Current mode 1: Find the maximum alpha value")
                            alpha = '%.6f' % (0.0001390 + flag_1 / 1000000)  # find best alpha
                            alpha = float(alpha)
                            print('dsfr feature selection, alpha = {}'.format(alpha))
                            inner_dsfr_feature_choose, out_dsfr_feature_choose, dsfrfl = \
                                self.choose_feature(dsfrfeature, feature_out=dsfrfeature_out, \
                                                    savepath=final_savepath + r"/dsfr", uselasso=True, alpha=alpha,
                                                    radio_dsfr='dsfr')
                        elif choose_alpha_type == 2:  # Fixed alpha, also only select once
                            if flag_1 == 1:
                                alpha = 0.0021  # fixed
                                print("Current mode 2: Fixed alpha to " + str(alpha))
                                inner_dsfr_feature_choose, out_dsfr_feature_choose, dsfrfl = \
                                    self.choose_feature(dsfrfeature, feature_out=dsfrfeature_out, \
                                                        savepath=final_savepath + r"/dsfr", uselasso=True, alpha=alpha)

                        else:
                            raise Exception("The Choose_alpha_type must be set to '1' or '2' or '3' ")

                # If no features are selected, end the loop
                if radioflcv1 == 0 or dsfrfl == 0 or tfnfl == 0:
                    print('There are 0 features remaining after filtering!')
                else:
                    # Build containers for storing results
                    label_CV = []
                    id_CV = []

                    # lr
                    dsfrprob_lr = []
                    dsfrprob_lr_out = []
                    radioprob_lr = []
                    radioprob_lr_out = []

                    # svm
                    dsfrprob_svm = []
                    dsfrprob_svm_out = []
                    radioprob_svm = []
                    radioprob_svm_out = []

                    # RF
                    dsfrprob_rf = []
                    dsfrprob_rf_out = []
                    radioprob_rf = []
                    radioprob_rf_out = []

                    fold_flag = 1  # Record which fold of cross-validation is running.
                    smo = SMOTE()  # Repeat sampling of data, unbalanced positive and negative sample ratios
                    smosvm = SVMSMOTE()
                    # ====================================== Splitting experiment =============================================
                    for train, validation in sfolder.split(inner_radio_feature_choose, self.label):

                        print('Doing with fold: ', fold_flag, validation)
                        fold_flag = fold_flag + 1
                        if show_type == 3:
                            # Split data 0
                            # dsfr training set
                            # todo
                            dsfrfeature_train_1, label_train_dsfr = [inner_dsfr_feature_choose[train],
                                                                     self.label[train]]
                            # Omics training set
                            radiofeature_train_1, label_train_radio = [inner_radio_feature_choose[train],
                                                                       self.label[train]]
                            if (self.Smo):
                                dsfrfeature_train_1, label_train_dsfr = smo.fit_resample(dsfrfeature_train_1,
                                                                                         label_train_dsfr)
                                radiofeature_train_1, label_train_radio = smo.fit_resample(radiofeature_train_1,
                                                                                           label_train_radio)
                            if self.SvmSmo:
                                dsfrfeature_train_1, label_train_dsfr = smosvm.fit_resample(dsfrfeature_train_1,
                                                                                            label_train_dsfr)
                                radiofeature_train_1, label_train_radio = smosvm.fit_resample(radiofeature_train_1,
                                                                                              label_train_radio)

                            # dsfr validation set
                            dsfrfeature_vali_1, label_vali_dsfr = [inner_dsfr_feature_choose[validation],
                                                                   self.label[validation]]
                            # Omics validation set
                            radiofeature_vali_1, label_vali_radio = [inner_radio_feature_choose[validation],
                                                                     self.label[validation]]
                            if self.Smo:
                                dsfrfeature_vali_1, label_vali_dsfr = smo.fit_resample(dsfrfeature_vali_1,
                                                                                       label_vali_dsfr)
                                radiofeature_vali_1, label_vali_radio = smo.fit_resample(radiofeature_vali_1,
                                                                                         label_vali_radio)
                            if self.SvmSmo:
                                dsfrfeature_vali_1, label_vali_dsfr = smosvm.fit_resample(dsfrfeature_vali_1,
                                                                                          label_vali_dsfr)
                                radiofeature_vali_1, label_vali_radio = smosvm.fit_resample(radiofeature_vali_1,
                                                                                            label_vali_radio)

                            id_CV = id_CV + list(id_all[validation])  # Validation set patient id

                            # Test set is copied, just change the name
                            if self.test_outside:
                                dsfrfeature_out_test = copy.deepcopy(out_dsfr_feature_choose)
                                radiofeature_out_test = copy.deepcopy(out_radio_feature_choose)

                        if show_type == 3:
                            # Standardize semantic features
                            StandardScaler = preprocessing.StandardScaler()  # Initialize standardization parameters
                            dsfrfeature_train = StandardScaler.fit_transform(dsfrfeature_train_1)  # Fit
                            dsfrfeature_vali = StandardScaler.transform(
                                dsfrfeature_vali_1)  # Standardize validation set according to training set
                            # Standardize omics features
                            StandardScaler2 = preprocessing.StandardScaler()  # Initialize standardization parameters
                            radiofeature_train = StandardScaler2.fit_transform(radiofeature_train_1)  # Fit
                            radiofeature_vali = StandardScaler2.transform(
                                radiofeature_vali_1)  # Standardize validation set according to training set
                            # Standardize test set
                            if self.test_outside:
                                dsfrfeature_out_test = StandardScaler.transform(
                                    dsfrfeature_out_test)  # Standardize semantic features
                                radiofeature_out_test = StandardScaler2.transform(
                                    radiofeature_out_test)  # Standardize omics features

                        # ====================================== Classifier settings =============================================
                        self.class_para()  # Load parameters
                        # Logistic regression classifier
                        if self.UseLR:
                            print('LR beginning!  ', end="")
                            time_1 = time.time()  # Time LR
                            clf = GridSearchCV(LogisticRegression(solver='lbfgs', random_state=grid_rdstate),
                                               # ,class_weight='balanced'),
                                               self.lr_grid_search_param, cv=3, scoring='roc_auc')
                            if show_type == 3:
                                # dsfr feature results
                                if 'dsfr' in self.data_name:
                                    clf.fit(dsfrfeature_train, label_train_dsfr)  # Fit data SMOTE
                                    save_bestcoff(save_path=output_file_path,
                                                  bestcoff=clf.best_estimator_.coef_,
                                                  type='dsfr')
                                    # Predict validation set
                                    y_prob = clf.predict_proba(
                                        dsfrfeature_vali)  # Validation set results, output probability scores
                                    dsfrprob_lr = dsfrprob_lr + list(y_prob[:, 1])  # Save probability scores
                                    # Test external validation
                                    if self.test_outside:
                                        y_prob = clf.predict_proba(
                                            dsfrfeature_out_test)  # Test set results, output probability scores
                                        dsfrprob_lr_out.append(list(y_prob[:, 1]))  # Save probability scores
                                    self.save_param(os.path.join(final_savepath, 'dsfr_param.xlsx'), flag_1, clf,
                                                    'LR')  # Save parameters

                                # Omics feature results
                                if 'radio' in self.data_name:
                                    clf.fit(radiofeature_train, label_train_radio)  # Fit data SMOTE
                                    save_bestcoff(save_path=output_file_path,
                                                  bestcoff=clf.best_estimator_.coef_,
                                                  type='radio')
                                    # Predict validation set
                                    y_prob = clf.predict_proba(radiofeature_vali)
                                    radioprob_lr = radioprob_lr + list(y_prob[:, 1])
                                    if self.test_outside:
                                        y_prob = clf.predict_proba(radiofeature_out_test)
                                        radioprob_lr_out.append(list(y_prob[:, 1]))
                                    self.save_param(os.path.join(final_savepath, 'radio_param.xlsx'), flag_1, clf,
                                                    'LR')  # Save parameters

                            time_2 = time.time()
                            print('It spent  ' + str(time_2 - time_1) + ' s')

                        # Support vector machine classifier
                        if self.UseSVM:
                            print('SVM beginning!  ', end="")
                            time_3 = time.time()  # Time SVM
                            clf = GridSearchCV(SVC(probability=True, random_state=grid_rdstate),
                                               self.svm_grid_search_param,
                                               cv=3, scoring='roc_auc')

                            # dsfr feature results
                            if 'dsfr' in self.data_name:
                                clf.fit(dsfrfeature_train, label_train_dsfr)  # Fit data SMOTE
                                # Predict validation set
                                y_prob = clf.predict_proba(
                                    dsfrfeature_vali)  # Validation set results, output probability scores
                                dsfrprob_svm = dsfrprob_svm + list(y_prob[:, 1])  # Save probability scores
                                # Test external validation
                                if self.test_outside:
                                    y_prob = clf.predict_proba(
                                        dsfrfeature_out_test)  # Test set results, output probability scores
                                    dsfrprob_svm_out.append(list(y_prob[:, 1]))  # Save probability scores
                                self.save_param(os.path.join(final_savepath, 'dsfr_param.xlsx'), flag_1, clf,
                                                'SVM')  # Save parameters

                            # Omics feature results
                            if 'radio' in self.data_name:
                                clf.fit(radiofeature_train, label_train_radio)  # SMOTE
                                # Predict validation set
                                y_prob = clf.predict_proba(radiofeature_vali)
                                radioprob_svm = radioprob_svm + list(y_prob[:, 1])
                                if self.test_outside:
                                    y_prob = clf.predict_proba(radiofeature_out_test)
                                    radioprob_svm_out.append(list(y_prob[:, 1]))
                                self.save_param(os.path.join(final_savepath, 'radio_param.xlsx'), flag_1, clf,
                                                'SVM')  # Save parameters

                            time_4 = time.time()
                            print('It spent  ' + str(time_4 - time_3) + ' s')

                        # Random forest classifier
                        if self.UseRF:
                            print('RF beginning!  ', end="")
                            time_5 = time.time()  # Time RF
                            clf = GridSearchCV(RandomForestClassifier(random_state=grid_rdstate),
                                               self.rf_grid_search_param,
                                               cv=3, scoring='roc_auc')

                            # dsfr feature results
                            if 'dsfr' in self.data_name:
                                clf.fit(dsfrfeature_train, label_train_dsfr)  # Fit data
                                # Predict validation set
                                y_prob = clf.predict_proba(
                                    dsfrfeature_vali)  # Validation set results, output probability scores
                                dsfrprob_rf = dsfrprob_rf + list(y_prob[:, 1])  # Save probability scores
                                # Test external validation
                                if self.test_outside:
                                    y_prob = clf.predict_proba(
                                        dsfrfeature_out_test)  # Test set results, output probability scores
                                    dsfrprob_rf_out.append(list(y_prob[:, 1]))  # Save probability scores
                                self.save_param(os.path.join(final_savepath, 'dsfr_param.xlsx'), flag_1, clf,
                                                'RF')  # Save parameters

                            # Omics feature results
                            if 'radio' in self.data_name:
                                clf.fit(radiofeature_train, label_train_radio)
                                # Predict validation set
                                y_prob = clf.predict_proba(radiofeature_vali)
                                radioprob_rf = radioprob_rf + list(y_prob[:, 1])
                                if self.test_outside:
                                    y_prob = clf.predict_proba(radiofeature_out_test)
                                    radioprob_rf_out.append(list(y_prob[:, 1]))
                                self.save_param(os.path.join(final_savepath, 'radio_param.xlsx'), flag_1, clf,
                                                'RF')  # Save parameters

                            time_6 = time.time()
                            print('It spent  ' + str(time_6 - time_5) + ' s')

                        label_CV = label_CV + list(label_vali_radio)  # Validation set label smote

                    final_label_CV.append(label_CV)
                    print('ID', id_CV)
                    print('ID', label_CV)

                    # ====================================== Save results =============================================
                    # Calculate scores
                    if show_type == 1:  # Single dsfr result
                        prob_svm = dsfrprob_svm
                        prob_lr = dsfrprob_lr
                        prob_rf = dsfrprob_rf
                        prob_svm_final.append(prob_svm)
                        prob_lr_final.append(prob_lr)
                        prob_rf_final.append(prob_rf)
                        if self.test_outside:
                            prob_svm_out = dsfrprob_svm_out
                            prob_lr_out = dsfrprob_lr_out
                            prob_rf_out = dsfrprob_rf_out
                            prob_svm_out_final.append(prob_svm_out)
                            prob_lr_out_final.append(prob_lr_out)
                            prob_rf_out_final.append(prob_rf_out)
                    elif show_type == 2:  # Single omics result
                        prob_svm = radioprob_svm
                        prob_lr = radioprob_lr
                        prob_rf = radioprob_rf
                        prob_svm_final.append(prob_svm)
                        prob_lr_final.append(prob_lr)
                        prob_rf_final.append(prob_rf)
                        if self.test_outside:
                            prob_svm_out = radioprob_svm_out
                            prob_lr_out = radioprob_lr_out
                            prob_rf_out = radioprob_rf_out
                            prob_svm_out_final.append(prob_svm_out)
                            prob_lr_out_final.append(prob_lr_out)
                            prob_rf_out_final.append(prob_rf_out)
                    elif show_type == 3:  # Add dsfr and omics scores
                        prob_svm = (np.array(dsfrprob_svm) + np.array(radioprob_svm)) / 2
                        prob_lr = (np.array(dsfrprob_lr) + np.array(radioprob_lr)) / 2
                        prob_rf = (np.array(dsfrprob_rf) + np.array(radioprob_rf)) / 2
                        prob_svm_final.append(prob_svm)
                        prob_lr_final.append(prob_lr)
                        prob_rf_final.append(prob_rf)
                        if self.test_outside:
                            prob_svm_out = (np.array(dsfrprob_svm_out) + np.array(radioprob_svm_out)) / 2
                            prob_lr_out = (np.array(dsfrprob_lr_out) + np.array(radioprob_lr_out)) / 2
                            prob_rf_out = (np.array(dsfrprob_rf_out) + np.array(radioprob_rf_out)) / 2
                            prob_svm_out_final.append(prob_svm_out)
                            prob_lr_out_final.append(prob_lr_out)
                            prob_rf_out_final.append(prob_rf_out)

                    # Calculate internal validation set data metrics
                    if self.UseSVM:
                        acc_svm, sen_svm, spc_svm, AUC_svm, bst_svm = self.getAccSenSpcAuc(label_CV, prob_svm)
                    if self.UseLR:
                        acc_lr, sen_lr, spc_lr, AUC_lr, bst_lr = self.getAccSenSpcAuc(label_CV, prob_lr)
                    if self.UseRF:
                        acc_rf, sen_rf, spc_rf, AUC_rf, bst_rf = self.getAccSenSpcAuc(label_CV, prob_rf)

                    # Calculate external validation metrics. Note that when calculating metrics, use the best threshold of internal data!!!
                    if self.test_outside:
                        label_CV_out = list(self.label_out)
                        if self.UseSVM:
                            prob_svm_out = list(np.mean(np.array(prob_svm_out), axis=0))
                            acc_svm_out, sen_svm_out, spc_svm_out, AUC_svm_out, bst_svm_out = self.getAccSenSpcAuc(
                                label_CV_out, prob_svm_out, pre_bestthresold=bst_svm)

                        if self.UseLR:
                            prob_lr_out = list(np.mean(np.array(prob_lr_out), axis=0))
                            acc_lr_out, sen_lr_out, spc_lr_out, AUC_lr_out, bst_lr_out = self.getAccSenSpcAuc(
                                label_CV_out, prob_lr_out, pre_bestthresold=bst_lr)

                        if self.UseRF:
                            prob_rf_out = list(np.mean(np.array(prob_rf_out), axis=0))
                            acc_rf_out, sen_rf_out, spc_rf_out, AUC_rf_out, bst_rf_out = self.getAccSenSpcAuc(
                                label_CV_out, prob_rf_out, pre_bestthresold=bst_rf)

                    # ====================================== 打印结果 =============================================
                    # 内部结果
                    writewords = str(flag_1) + r'===========================================================' + '\n' + \
                                 r' $  holdout: ' + str(self.holdout) + ', distill_mode : ' + self.distill_mode + ',' + \
                                 '  cvrdstate: ' + str(cvrdstate) + \
                                 '  grid_rdstate: ' + str(grid_rdstate) + '\n' + \
                                 '  k_fold:' + str(self.k_fold) + '\n'
                    if self.UseSVM:
                        writewords = writewords + r' @ svm  @ AUC: ' + str('%03f' % AUC_svm) + \
                                     ',Acc: ' + str('%03f' % acc_svm) + \
                                     ',Sen: ' + str('%03f' % sen_svm) + \
                                     ',Spc: ' + str('%03f' % spc_svm) + \
                                     ',Best_thresold: ' + str('%03f' % bst_svm) + '\n'
                    if self.UseLR:
                        writewords = writewords + r' @ lr   @ AUC: ' + str('%03f' % AUC_lr) + \
                                     ',Acc: ' + str('%03f' % acc_lr) + \
                                     ',Sen: ' + str('%03f' % sen_lr) + \
                                     ',Spc: ' + str('%03f' % spc_lr) + \
                                     ',Best_thresold: ' + str('%03f' % bst_lr) + '\n'
                    if self.UseRF:
                        writewords = writewords + r' @ RF   @ AUC: ' + str('%03f' % AUC_rf) + \
                                     ',Acc: ' + str('%03f' % acc_rf) + \
                                     ',Sen: ' + str('%03f' % sen_rf) + \
                                     ',Spc: ' + str('%03f' % spc_rf) + \
                                     ',Best_thresold: ' + str('%03f' % bst_rf) + '\n'
                    print(writewords)

                    # 外部结果
                    writewords = '\n' + r'* outside *' + '\n'
                    if self.UseSVM:
                        writewords = writewords + r' @ svm  @ outAUC: ' + str('%03f' % AUC_svm_out) + \
                                     ',Acc: ' + str('%03f' % acc_svm_out) + \
                                     ',Sen: ' + str('%03f' % sen_svm_out) + \
                                     ',Spc: ' + str('%03f' % spc_svm_out) + \
                                     ',Best_thresold: ' + str('%03f' % bst_svm) + '\n'
                    if self.UseLR:
                        writewords = writewords + r' @ lr   @ outAUC: ' + str('%03f' % AUC_lr_out) + \
                                     ',Acc: ' + str('%03f' % acc_lr_out) + \
                                     ',Sen: ' + str('%03f' % sen_lr_out) + \
                                     ',Spc: ' + str('%03f' % spc_lr_out) + \
                                     ',Best_thresold: ' + str('%03f' % bst_lr) + '\n'
                    if self.UseRF:
                        writewords = writewords + r' @ RF   @ outAUC: ' + str('%03f' % AUC_rf_out) + \
                                     ',Acc: ' + str('%03f' % acc_rf_out) + \
                                     ',Sen: ' + str('%03f' % sen_rf_out) + \
                                     ',Spc: ' + str('%03f' % spc_rf_out) + \
                                     ',Best_thresold: ' + str('%03f' % bst_rf_out) + '\n'
                    print(writewords)

                    # ====================================== Print results =============================================
                    # Internal results
                    writewords = str(flag_1) + r'===========================================================' + '\n' + \
                                 r' $  holdout: ' + str(self.holdout) + ', distill_mode : ' + self.distill_mode + ',' + \
                                 '  cvrdstate: ' + str(cvrdstate) + \
                                 '  grid_rdstate: ' + str(grid_rdstate) + '\n' + \
                                 '  k_fold:' + str(self.k_fold) + '\n'
                    if self.UseSVM:
                        writewords = writewords + r' @ svm  @ AUC: ' + str('%03f' % AUC_svm) + \
                                     ',Acc: ' + str('%03f' % acc_svm) + \
                                     ',Sen: ' + str('%03f' % sen_svm) + \
                                     ',Spc: ' + str('%03f' % spc_svm) + \
                                     ',Best_thresold: ' + str('%03f' % bst_svm) + '\n'
                    if self.UseLR:
                        writewords = writewords + r' @ lr   @ AUC: ' + str('%03f' % AUC_lr) + \
                                     ',Acc: ' + str('%03f' % acc_lr) + \
                                     ',Sen: ' + str('%03f' % sen_lr) + \
                                     ',Spc: ' + str('%03f' % spc_lr) + \
                                     ',Best_thresold: ' + str('%03f' % bst_lr) + '\n'
                    if self.UseRF:
                        writewords = writewords + r' @ RF   @ AUC: ' + str('%03f' % AUC_rf) + \
                                     ',Acc: ' + str('%03f' % acc_rf) + \
                                     ',Sen: ' + str('%03f' % sen_rf) + \
                                     ',Spc: ' + str('%03f' % spc_rf) + \
                                     ',Best_thresold: ' + str('%03f' % bst_rf) + '\n'
                    print(writewords)

                    # External results
                    writewords = '\n' + r'* outside *' + '\n'
                    if self.UseSVM:
                        writewords = writewords + r' @ svm  @ outAUC: ' + str('%03f' % AUC_svm_out) + \
                                     ',Acc: ' + str('%03f' % acc_svm_out) + \
                                     ',Sen: ' + str('%03f' % sen_svm_out) + \
                                     ',Spc: ' + str('%03f' % spc_svm_out) + \
                                     ',Best_thresold: ' + str('%03f' % bst_svm) + '\n'
                    if self.UseLR:
                        writewords = writewords + r' @ lr   @ outAUC: ' + str('%03f' % AUC_lr_out) + \
                                     ',Acc: ' + str('%03f' % acc_lr_out) + \
                                     ',Sen: ' + str('%03f' % sen_lr_out) + \
                                     ',Spc: ' + str('%03f' % spc_lr_out) + \
                                     ',Best_thresold: ' + str('%03f' % bst_lr) + '\n'
                    if self.UseRF:
                        writewords = writewords + r' @ RF   @ outAUC: ' + str('%03f' % AUC_rf_out) + \
                                     ',Acc: ' + str('%03f' % acc_rf_out) + \
                                     ',Sen: ' + str('%03f' % sen_rf_out) + \
                                     ',Spc: ' + str('%03f' % spc_rf_out) + \
                                     ',Best_thresold: ' + str('%03f' % bst_rf_out) + '\n'
                    print(writewords)

                    # ====================================== Save results =============================================
                    # Save all internal (validation set) and test set results in excel
                    filename = final_savepath + sep + 'all_results.xlsx'
                    self.writesimple(filename, 'turn_alpha', 1, 1)  # Row 1, column 1 'Round and alpha'
                    self.writesimple(filename, 'holdout', 1, 2)  # Row 1, column 2 'holdout'
                    self.writesimple(filename, 'distill_mode', 1, 3)  # Row 1, column 3 'distill_mode'
                    self.writesimple(filename, 'cvrdstate', 1, 4)  # Row 1, column 4 'cvrdstate'
                    self.writesimple(filename, 'grid_rdstate', 1, 5)  # Row 1, column 5 'grid_rdstate'
                    self.writesimple(filename, 'k_fold', 1, 6)  # Row 1, column 6 'k_fold'

                    self.writesimple(filename, str(flag_1) + '_' + str(alpha), flag_1 + 1, 1)  # Round
                    self.writesimple(filename, str(self.holdout), flag_1 + 1, 2)
                    self.writesimple(filename, str(self.distill_mode), flag_1 + 1, 3)
                    self.writesimple(filename, str(cvrdstate), flag_1 + 1, 4)
                    self.writesimple(filename, str(grid_rdstate), flag_1 + 1, 5)
                    self.writesimple(filename, str(self.k_fold), flag_1 + 1, 6)
                    if self.UseLR:
                        # Internal
                        self.writesimple(filename, 'LR', 1, 7)
                        self.writesimple(filename, 'In_AUC', 1, 8)
                        self.writesimple(filename, 'In_Acc', 1, 9)
                        self.writesimple(filename, 'In_Sen', 1, 10)
                        self.writesimple(filename, 'In_Spc', 1, 11)
                        self.writesimple(filename, 'In_Best_thresold', 1, 12)
                        self.writesimple(filename, str('%03f' % AUC_lr), flag_1 + 1, 8)
                        self.writesimple(filename, str('%03f' % acc_lr), flag_1 + 1, 9)
                        self.writesimple(filename, str('%03f' % sen_lr), flag_1 + 1, 10)
                        self.writesimple(filename, str('%03f' % spc_lr), flag_1 + 1, 11)
                        self.writesimple(filename, str('%03f' % bst_lr), flag_1 + 1, 12)

                        # External
                        if self.test_outside:
                            self.writesimple(filename, 'Out_AUC', 1, 13)
                            self.writesimple(filename, 'Out_Acc', 1, 14)
                            self.writesimple(filename, 'Out_Sen', 1, 15)
                            self.writesimple(filename, 'Out_Spc', 1, 16)
                            self.writesimple(filename, 'Out_Best_thresold', 1, 17)
                            self.writesimple(filename, str('%03f' % AUC_lr_out), flag_1 + 1, 13)
                            self.writesimple(filename, str('%03f' % acc_lr_out), flag_1 + 1, 14)
                            self.writesimple(filename, str('%03f' % sen_lr_out), flag_1 + 1, 15)
                            self.writesimple(filename, str('%03f' % spc_lr_out), flag_1 + 1, 16)
                            self.writesimple(filename, str('%03f' % bst_lr_out), flag_1 + 1, 17)
                    if self.UseSVM:
                        # Internal
                        self.writesimple(filename, 'SVM', 1, 18)
                        self.writesimple(filename, 'In_AUC', 1, 19)
                        self.writesimple(filename, 'In_Acc', 1, 20)
                        self.writesimple(filename, 'In_Sen', 1, 21)
                        self.writesimple(filename, 'In_Spc', 1, 22)
                        self.writesimple(filename, 'In_Best_thresold', 1, 23)
                        self.writesimple(filename, str('%03f' % AUC_svm), flag_1 + 1, 19)
                        self.writesimple(filename, str('%03f' % acc_svm), flag_1 + 1, 20)
                        self.writesimple(filename, str('%03f' % sen_svm), flag_1 + 1, 21)
                        self.writesimple(filename, str('%03f' % spc_svm), flag_1 + 1, 22)
                        self.writesimple(filename, str('%03f' % bst_svm), flag_1 + 1, 23)
                        # External
                        if self.test_outside:
                            self.writesimple(filename, 'Out_AUC', 1, 24)
                            self.writesimple(filename, 'Out_Acc', 1, 25)
                            self.writesimple(filename, 'Out_Sen', 1, 26)
                            self.writesimple(filename, 'Out_Spc', 1, 27)
                            self.writesimple(filename, 'Out_Best_thresold', 1, 28)
                            self.writesimple(filename, str('%03f' % AUC_svm_out), flag_1 + 1, 24)
                            self.writesimple(filename, str('%03f' % acc_svm_out), flag_1 + 1, 25)
                            self.writesimple(filename, str('%03f' % sen_svm_out), flag_1 + 1, 26)
                            self.writesimple(filename, str('%03f' % spc_svm_out), flag_1 + 1, 27)
                            self.writesimple(filename, str('%03f' % bst_svm_out), flag_1 + 1, 28)
                    if self.UseRF:
                        # Internal
                        self.writesimple(filename, 'RF', 1, 29)
                        self.writesimple(filename, 'In_AUC', 1, 30)
                        self.writesimple(filename, 'In_Acc', 1, 31)
                        self.writesimple(filename, 'In_Sen', 1, 32)
                        self.writesimple(filename, 'In_Spc', 1, 33)
                        self.writesimple(filename, 'In_Best_thresold', 1, 34)
                        self.writesimple(filename, str('%03f' % AUC_rf), flag_1 + 1, 30)
                        self.writesimple(filename, str('%03f' % acc_rf), flag_1 + 1, 31)
                        self.writesimple(filename, str('%03f' % sen_rf), flag_1 + 1, 32)
                        self.writesimple(filename, str('%03f' % spc_rf), flag_1 + 1, 33)
                        self.writesimple(filename, str('%03f' % bst_rf), flag_1 + 1, 34)
                        # External
                        if self.test_outside:
                            self.writesimple(filename, 'Out_AUC', 1, 35)
                            self.writesimple(filename, 'Out_Acc', 1, 36)
                            self.writesimple(filename, 'Out_Sen', 1, 37)
                            self.writesimple(filename, 'Out_Spc', 1, 38)
                            self.writesimple(filename, 'Out_Best_thresold', 1, 39)
                            self.writesimple(filename, str('%03f' % AUC_rf_out), flag_1 + 1, 35)
                            self.writesimple(filename, str('%03f' % acc_rf_out), flag_1 + 1, 36)
                            self.writesimple(filename, str('%03f' % sen_rf_out), flag_1 + 1, 37)
                            self.writesimple(filename, str('%03f' % spc_rf_out), flag_1 + 1, 38)
                            self.writesimple(filename, str('%03f' % bst_rf_out), flag_1 + 1, 39)

                    # ====================================== Save scores =============================================
                    # Save scores of each round of validation set in excel
                    filename = final_savepath + sep + 'In_prob_results.xlsx'
                    if self.UseLR:
                        self.writesimple(filename, 'no' + str(flag_1) + 'turn', 1, flag_1 * 4 - 3, sheetname='LR')
                        self.writesimple(filename, 'ID', 1, flag_1 * 4 - 2, sheetname='LR')
                        self.writesimple(filename, 'Prob', 1, flag_1 * 4 - 1, sheetname='LR')
                        self.writesimple(filename, 'label', 1, flag_1 * 4, sheetname='LR')
                        for i in range(len(id_CV)):
                            self.writesimple(filename, str(id_CV[i]), i + 2, flag_1 * 4 - 2, sheetname='LR')  # Save ID
                            self.writesimple(filename, str(prob_lr[i]), i + 2, flag_1 * 4 - 1,
                                             sheetname='LR')  # Save score
                            self.writesimple(filename, str(label_CV[i]), i + 2, flag_1 * 4,
                                             sheetname='LR')  # Save label
                    if self.UseSVM:
                        self.writesimple(filename, 'no' + str(flag_1) + 'turn', 1, flag_1 * 4 - 3, sheetname='SVM')
                        self.writesimple(filename, 'ID', 1, flag_1 * 4 - 2, sheetname='SVM')
                        self.writesimple(filename, 'Prob', 1, flag_1 * 4 - 1, sheetname='SVM')
                        self.writesimple(filename, 'label', 1, flag_1 * 4, sheetname='SVM')
                        for i in range(len(id_CV)):
                            self.writesimple(filename, str(id_CV[i]), i + 2, flag_1 * 4 - 2, sheetname='SVM')  # Save ID
                            self.writesimple(filename, str(prob_svm[i]), i + 2, flag_1 * 4 - 1,
                                             sheetname='SVM')  # Save score
                            self.writesimple(filename, str(label_CV[i]), i + 2, flag_1 * 4,
                                             sheetname='SVM')  # Save label
                    if self.UseRF:
                        self.writesimple(filename, 'no' + str(flag_1) + 'turn', 1, flag_1 * 4 - 3, sheetname='RF')
                        self.writesimple(filename, 'ID', 1, flag_1 * 4 - 2, sheetname='RF')
                        self.writesimple(filename, 'Prob', 1, flag_1 * 4 - 1, sheetname='RF')
                        self.writesimple(filename, 'label', 1, flag_1 * 4, sheetname='RF')
                        for i in range(len(id_CV)):
                            self.writesimple(filename, str(id_CV[i]), i + 2, flag_1 * 4 - 2, sheetname='RF')  # Save ID
                            self.writesimple(filename, str(prob_rf[i]), i + 2, flag_1 * 4 - 1,
                                             sheetname='RF')  # Save score
                            self.writesimple(filename, str(int(label_CV[i])), i + 2, flag_1 * 4,
                                             sheetname='RF')  # Save label
                    # Save the written excel
                    source_file = filename
                    target_file = final_savepath + sep + 'In_prob_results2.xlsx'
                    shutil.copyfile(source_file, target_file)

                    # Save scores of each round of external test set in excel
                    if self.test_outside:
                        filename = final_savepath + sep + 'Out_prob_results.xlsx'
                        if self.UseLR:
                            self.writesimple(filename, 'no' + str(flag_1) + 'turn', 1, flag_1 * 4 - 3, sheetname='LR')
                            self.writesimple(filename, 'ID', 1, flag_1 * 4 - 2, sheetname='LR')
                            self.writesimple(filename, 'Prob', 1, flag_1 * 4 - 1, sheetname='LR')
                            self.writesimple(filename, 'label', 1, flag_1 * 4, sheetname='LR')
                            for j in range(len(prob_lr_out) - len(id_all_out)):
                                id_all_out.append('{}_0000'.format(j))
                            for i in range(len(prob_lr_out)):
                                self.writesimple(filename, str(id_all_out[i]), i + 2, flag_1 * 4 - 2, sheetname='LR')  # Save ID
                                self.writesimple(filename, str(prob_lr_out[i]), i + 2, flag_1 * 4 - 1, sheetname='LR')  # Save score
                                self.writesimple(filename, str(label_CV_out[i]), i + 2, flag_1 * 4, sheetname='LR')  # Save label
                        if self.UseSVM:
                            self.writesimple(filename, 'no' + str(flag_1) + 'turn', 1, flag_1 * 4 - 3, sheetname='SVM')
                            self.writesimple(filename, 'ID', 1, flag_1 * 4 - 2, sheetname='SVM')
                            self.writesimple(filename, 'Prob', 1, flag_1 * 4 - 1, sheetname='SVM')
                            self.writesimple(filename, 'label', 1, flag_1 * 4, sheetname='SVM')
                            for j in range(len(prob_svm_out) - len(id_all_out)):
                                id_all_out.append('{}_0000'.format(j))
                            for i in range(len(prob_svm_out)):
                                self.writesimple(filename, str(id_all_out[i]), i + 2, flag_1 * 4 - 2,
                                                 sheetname='SVM')  # Save ID
                                self.writesimple(filename, str(prob_svm_out[i]), i + 2, flag_1 * 4 - 1,
                                                 sheetname='SVM')  # Save score
                                self.writesimple(filename, str(int(label_CV_out[i])), i + 2, flag_1 * 4,
                                                 sheetname='SVM')  # Save label
                        if self.UseRF:
                            self.writesimple(filename, 'no' + str(flag_1) + 'turn', 1, flag_1 * 4 - 3, sheetname='RF')
                            self.writesimple(filename, 'ID', 1, flag_1 * 4 - 2, sheetname='RF')
                            self.writesimple(filename, 'Prob', 1, flag_1 * 4 - 1, sheetname='RF')
                            self.writesimple(filename, 'label', 1, flag_1 * 4, sheetname='RF')
                            for j in range(len(prob_rf_out) - len(id_all_out)):
                                id_all_out.append('{}_0000'.format(j))
                            for i in range(len(prob_rf_out)):
                                self.writesimple(filename, str(int(id_all_out[i])), i + 2, flag_1 * 4 - 2,
                                                 sheetname='RF')  # Save ID
                                self.writesimple(filename, str(prob_rf_out[i]), i + 2, flag_1 * 4 - 1,
                                                 sheetname='RF')  # Save score
                                self.writesimple(filename, str(int(label_CV_out[i])), i + 2, flag_1 * 4,
                                                 sheetname='RF')  # Save label
                    # Save the written excel
                    source_file = filename
                    target_file = final_savepath + sep + 'Out_prob_results2.xlsx'
                    shutil.copyfile(source_file, target_file)
if __name__=="__main__":
# ====================================== Path addresses =============================================
    pkl_file = open(r'/data/test.pkl', 'rb')  # pkl file obtained from Excel_to_pkl of preprocessing
    savepath = os.path.join(r'/data/results/')  # Result save path
    output_file_path = savepath + "model_weights.txt" # File to record model weights

    other.judgedir(savepath)
    print('savepath:', savepath)

    task = '6'  # Task type
    # Input features
    radio_feature_path_6 = r'/data/radio_feature/'  # VAT or SAT
    dsfr_feature_path_6 = r'/data/dl_feature/'

    data_name = ['dsfr','radio'] # How many kinds of data are there

    Classifier(pkl_file=pkl_file,savepath=savepath,
               radio_feature_path_0=None,dsfr_feature_path_0=None,
              radio_feature_path_6=radio_feature_path_6,dsfr_feature_path_6=dsfr_feature_path_6,
               task=task,data_name=data_name).run(3,1,1)